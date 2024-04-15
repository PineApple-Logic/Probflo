import asyncio
from flask import Flask, render_template, request, jsonify, send_file,redirect
from flask_socketio import SocketIO, emit
from openpyxl import Workbook
import sqlite3
import json
import os
from hypercorn.config import Config
from hypercorn.asyncio import serve
import nodeszh
from nodegen import process_excel_file
import subprocess
from pathlib import Path
import cmd_out
import webbrowser
import discharge 

# script_dir = os.path.dirname(os.path.realpath(__file__))
# os.chdir()
current_directory = Path.cwd()
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'txt','xlsx'}
DATABASE = 'db.sqlite3'
socketio = SocketIO(app)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def create_tables(node_data):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS node_data (
            id INTEGER PRIMARY KEY,
            name TEXT,
            left TEXT,
            top TEXT,
            states TEXT,
            children TEXT,
            "values" TEXT,
            parents TEXT
        )
    ''')

    # Insert nodeData into the node_data table
    for node in node_data:
        cursor.execute('''
            INSERT INTO node_data (name, left, top, states, children, "values", parents)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            node.get('name', ''),
            node.get('left', ''),
            node.get('top', ''),
            json.dumps(node.get('states', [])),
            json.dumps(node.get('children', [])),
            json.dumps(node.get('values', [])),
            json.dumps(node.get('parents', [])),
        ))

    conn.commit()
    conn.close()


def initialize_data_from_database():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    # Fetch data from the node_data table
    cursor.execute('SELECT * FROM node_data')
    rows = cursor.fetchall()

    # Transform rows into a list of dictionaries (node_data)
    node_data = []
    for row in rows:
       
        node_data.append({
            "name": row[1],
            "left": row[2],
            "top": row[3],
            "states": json.loads(row[4]),
            "children": json.loads(row[5]),
            "values": json.loads(row[6]),
            "parents": json.loads(row[7])
        })

    conn.close()
    return node_data

# initialise from file 
def initialize_data():
    node_data = []
    node_data += nodeszh.nodeData
    return node_data


def build_parent_child_map(node_data):
    parent_child_map = {}

    # Populate parent_child_map with parent-child relationships
    for node in node_data:
        if "children" in node and node["children"]:
            for child_name in node["children"]:
                if child_name in parent_child_map:
                    parent_child_map[child_name].append(node["name"])
                else:
                    parent_child_map[child_name] = [node["name"]]

    return parent_child_map


def update_nodes_with_values_and_parents(node_data, parent_child_map):
    for node in node_data:
        node["values"] = [[], []]
        if node["name"] in parent_child_map:
            node["parents"] = parent_child_map[node["name"]]
        else:
            node["parents"] = []

    return node_data


def print_formatted_json(node_data):
    formatted_json = json.dumps(node_data, indent=4)
    # print(formatted_json)


def run_flask_app():
    @app.route("/")
    def hello_world():
        # return render_template("main.html")
        return redirect("/water")

    @app.route("/water")
    def water():
        return render_template("water.html")
    
    @app.route("/netica")
    def netica():
        network_dir = os.path.join(current_directory,"network")
        network_path = os.path.join(network_dir,"Balule.neta")
        casefile_path = os.path.join(current_directory,process_excel_file(os.path.join(app.config['UPLOAD_FOLDER'], 'netica_case.xlsx')))
        command = f'cmd /C "compute_neticajar.cmd \"{network_path}\"  \"{casefile_path}\""'
        print(command)
        output_file_path = 'dump.txt'
        output_data =None
        
        with open(output_file_path, 'w') as output_file:
          result = subprocess.run(command, shell=True, stdout=output_file)
        with open(output_file_path) as output_file:
            output_data = output_file.read()
        print(output_data)
        return render_template("netica.html",output_data=output_data,nodes = json.dumps(cmd_out.read_nodes_from_file()))

    @socketio.on('connect')
    def on_connect():
        print('A client connected')

    @socketio.on('disconnect')
    def on_disconnect():
        print('A client disconnected')

    @socketio.on('client_message')
    def handle_client_message(data):
        # print('Received from client:', data['message'])
        # Here you can process the data or send a response back to the client
        # For example:
        response_message = 'Message received on the server!'
        emit('server_message', {'message': response_message})

    @socketio.on('request_node_data')
    def send_node_data():
        emit('node_data', {'data': node_data})

    @socketio.on('update_node_data')
    def handle_update_node_data(data):
        updated_node_data = data['data']
        global node_data
        node_data = updated_node_data
        # Emit the updated data to all connected clients
        emit('node_data', {'data': node_data}, broadcast=True)


def run_hypercorn_server():
    port = int(os.environ.get('PORT', 4400))  # Default to 4400 if PORT environment variable not set
    config = Config()
  
    config.bind = [f"0.0.0.0:{port}"]
    asyncio.run(serve(app, config))


@app.route('/api/generate_cas_file', methods=['GET'])
def generate_cas():
    excel_file = os.path.join(app.config['UPLOAD_FOLDER'], 'netica_case.xlsx')
    
    cas_file_path = process_excel_file(excel_file)
    return send_file(cas_file_path, as_attachment=True, download_name='netica_case.cas', mimetype='application/octet-stream')
    


@app.route('/api/generate_excel', methods=['POST'])
def generate_excel():
        vals = None
        try:
            print("*"*10)
            string1 = request.form.get('string1', '')
            string2 = request.form.get('string2', '')
            print("*"*10)
            print(string1,string2)
            print("*"*10)
            natural_flows_file = request.files['naturalFlows']
            if natural_flows_file and allowed_file(natural_flows_file.filename):
            # Ensure the directory exists
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    # Save the file
                filename = os.path.join(app.config['UPLOAD_FOLDER'], 'naturalflow.xlsx')
                natural_flows_file.save(filename)
                vals = discharge.read_excel_file(filename)
                print("saved file")
            else:
               return jsonify({'error': 'Invalid file format or missing file'}), 400
              
            # Create a new Excel workbook
            workbook = Workbook()
            sheet = workbook.active
            nodes = string1.split(",")
            values = string2.replace("\"", "").replace(":", " ").split("||")

            # Place the strings in the first and second rows
            nodes += "DISCHARGE_YR++DISCHARGE_LF++DISCHARGE_HF++DISCHARGE_FD".split("++")
            # values += "{0 0.05, 0.9 0.2, 4.6 0.5, 20.9 0.2, 500.6 0.05}++{0 0.05, 0 0.2, 0 0.5, 0.9 0.2, 1.8 0.05}++{1.8 0.05, 4.6 0.2, 11.1 0.5, 20.9 0.2, 11.1 0.05}++{156.3 0.05, 233.1 0.2, 277.9 0.5, 438.2 0.2, 971.6 0.05}".split("++")
            values += vals.split("++")
            for i in range(len(nodes)):
                sheet.cell(row=1, column=i + 1, value=nodes[i])
                sheet.cell(row=2, column=i + 1, value=str(values[i]))

        #     # Save the workbook to a temporary file
            excel_file = os.path.join(app.config['UPLOAD_FOLDER'], 'netica_case.xlsx')
            workbook.save(excel_file)

        #     # Send the Excel file for download
            return send_file(excel_file, as_attachment=True, download_name='netica_case.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            return jsonify({'error': 'An error occurred while generating the Excel file'}), 500


if __name__ == "__main__":
    node_data = initialize_data() 
    create_tables(node_data)
    parent_child_map = build_parent_child_map(node_data)
    node_data = update_nodes_with_values_and_parents(node_data, parent_child_map)
    print_formatted_json(node_data)
    webbrowser.open('http://127.0.0.1:4400')  
    run_flask_app()
    run_hypercorn_server()
